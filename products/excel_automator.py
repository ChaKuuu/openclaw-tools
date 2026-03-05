#!/usr/bin/env python3
"""
Excel 自动化工具 - 立即可用
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
import os

class ExcelAutomator:
    """Excel 自动化工具"""
    
    def __init__(self, filename):
        self.filename = filename
        self.wb = None
        self.ws = None
    
    def create_report(self, data, headers, filename="report.xlsx"):
        """创建格式化报表"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # 写入表头
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filename)
        return filename
    
    def read_excel(self, sheet=0):
        """读取 Excel"""
        df = pd.read_excel(self.filename, sheet_name=sheet)
        return df
    
    def merge_files(self, files, output="merged.xlsx"):
        """合并多个 Excel 文件"""
        dfs = []
        for f in files:
            df = pd.read_excel(f)
            dfs.append(df)
        
        merged = pd.concat(dfs, ignore_index=True)
        merged.to_excel(output, index=False)
        return output

# 示例用法
if __name__ == "__main__":
    # 示例数据
    data = [
        ["产品A", 100, 50],
        ["产品B", 200, 80],
        ["产品C", 150, 60],
    ]
    headers = ["名称", "销量", "利润"]
    
    automator = ExcelAutomator("template.xlsx")
    output = automator.create_report(data, headers, "sales_report.xlsx")
    print(f"Created: {output}")
